from django.db.models import Q
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from .models import Batch
from django.db.models import ProtectedError
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.contrib import messages
from django.views import View
from django.shortcuts import get_object_or_404, redirect
from .forms import BatchForm, BatchUpdate
from django.urls import reverse_lazy
from categories.models import Category
from supplies.models import Supplier
from products.models import Product
from .models import Batch, BatchGroup
from django.core import serializers
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font
from datetime import datetime


class BatchList(ListView):
    model = Batch
    template_name = 'warehouse/products.html'
    context_object_name = 'batches'

    def get_queryset(self):
        # ... ваш существующий код без изменений ...
        queryset = Batch.objects.select_related(
            'product',
            'product__category',
            'product__supplier',
            'group'
        ).filter(quantity__gt=0).order_by('product__name', '-arrival_date')

        filters = Q()

        search_name = self.request.GET.get('search_name')
        if search_name:
            filters &= Q(product__name__icontains=search_name)

        category_id = self.request.GET.get('category')
        if category_id:
            filters &= Q(product__category_id=category_id)

        supplier_id = self.request.GET.get('supplier')
        if supplier_id:
            filters &= Q(product__supplier_id=supplier_id)

        if filters:
            queryset = queryset.filter(filters)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.all().order_by('name')
        context['suppliers'] = Supplier.objects.all().order_by('name')
        context['get_params'] = self.request.GET

        product_ids = list(set(Batch.objects.filter(quantity__gt=0).values_list('product_id', flat=True)))
        products_with_info = []

        for product_id in product_ids:
            try:
                product = Product.objects.select_related('category', 'supplier').get(id=product_id)
            except Product.DoesNotExist:
                continue

            product_batches = Batch.objects.filter(
                product_id=product_id,
                quantity__gt=0
            ).select_related('group')

            if product_batches.exists():
                total_quantity = sum(batch.quantity for batch in product_batches)
                last_batch = product_batches.order_by('-arrival_date').first()

                products_with_info.append({
                    'product': product,
                    'total_quantity': total_quantity,
                    'last_batch': last_batch,
                    'batches': product_batches
                })

        if self.request.GET.get('search_name'):
            search_name = self.request.GET.get('search_name')
            products_with_info = [p for p in products_with_info if search_name.lower() in p['product'].name.lower()]

        if self.request.GET.get('category'):
            category_id = int(self.request.GET.get('category'))
            products_with_info = [p for p in products_with_info if
                                  p['product'].category and p['product'].category.id == category_id]

        if self.request.GET.get('supplier'):
            supplier_id = int(self.request.GET.get('supplier'))
            products_with_info = [p for p in products_with_info if
                                  p['product'].supplier and p['product'].supplier.id == supplier_id]

        suppliers_dict = {}
        for product_info in products_with_info:
            supplier = product_info['product'].supplier
            supplier_name = supplier.name if supplier else 'Без поставщика'
            supplier_id = supplier.id if supplier else 'no_supplier'

            if supplier_id not in suppliers_dict:
                suppliers_dict[supplier_id] = {
                    'supplier': supplier,
                    'supplier_name': supplier_name,
                    'products': []
                }

            suppliers_dict[supplier_id]['products'].append(product_info)

        suppliers_list = []
        for supplier_id in sorted(suppliers_dict.keys(), key=lambda x: suppliers_dict[x]['supplier_name']):
            supplier_data = suppliers_dict[supplier_id]
            supplier_data['products'].sort(key=lambda x: x['product'].name)
            suppliers_list.append(supplier_data)

        context['suppliers_list'] = suppliers_list
        return context

class CreateBatch(CreateView):
    model = Batch
    form_class = BatchForm
    template_name = 'products/modal_form.html'
    success_url = reverse_lazy('warehouse:batch-list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['all_products'] = Product.objects.select_related('category', 'supplier').all().order_by('name')
        context['title'] = 'Провести новую поставку'
        context['action'] = 'Создать поставку'
        context['theme'] = 'green'
        return context

    def form_valid(self, form):
        batch = form.save(commit=False)
        group_batch = BatchGroup.objects.create(
            product = batch.product
        )
        batch.group = group_batch
        form.save()
        return super().form_valid(form)

class DeleteBatch(DeleteView):
    model = Batch
    template_name = 'products/modal_delete.html'
    success_url = reverse_lazy('warehouse:batch-list')
    context_object_name = 'batch'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            self.object.delete()
            messages.success(request, 'Партия успешно удалена!')
            return HttpResponseRedirect(reverse('warehouse:batch-list'))

        except ProtectedError as e:
            protected_objects = list(e.protected_objects)
            count = len(protected_objects)

            context = self.get_context_data(
                object=self.object,
                error=(
                    f'Невозможно удалить партию "{self.object}". '
                    f'Она связана с {count} продажами. '
                    f'Сначала удалите связанные продажи.'
                ),
                protected_objects=protected_objects
            )
            return self.render_to_response(context)


class ConsolidationBatch(View):
    def post(self, request, pk):
        product = get_object_or_404(Product, id=pk)

        batches = product.batches.filter(quantity__gt=0).order_by('arrival_date')

        if batches.exists():
            last_batch = batches.last()

            other_batches = batches.exclude(id=last_batch.id)
            total_quantity = sum(batch.quantity for batch in other_batches)

            if total_quantity > 0:
                last_batch.quantity += total_quantity
                last_batch.save()

                other_batches.update(quantity=0)

                messages.success(request, f'Поставки товара "{product.name}" объединены в последнюю партию!')
            else:
                messages.info(request, f'Нет поставок для объединения у товара "{product.name}"')
        else:
            messages.warning(request, f'У товара "{product.name}" нет поставок')
        return redirect('warehouse:batch-list')

class BatchUpdateInfo(UpdateView):
    model = Batch
    form_class = BatchUpdate
    template_name = 'products/modal_form.html'
    success_url = reverse_lazy('warehouse:batch-list')

    def form_valid(self, form):
        batch = self.object
        new_batch = Batch.objects.create(
            product=batch.product,
            group=batch.group,
            price=form.cleaned_data['price'],
            quantity=form.cleaned_data['quantity'],
            is_modified=True
        )

        batch.quantity = 0
        batch.save()

        messages.success(self.request, 'Партия успешно обновлена!')
        return super().form_valid(form)

def export_xlsx(request):
    batch_list_view = BatchList()
    batch_list_view.request = request
    queryset = batch_list_view.get_queryset()

    print(queryset)

    queryset = queryset.select_related('product', 'group')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Складские партии"

    headers = [
        'Название товара',
        'ID товара',
        'Цена',
        'Количество',
        'Дата поступления'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for row_num, batch in enumerate(queryset, 2):
        # Column 1: Product ID
        if batch.product:
            ws.cell(row=row_num, column=2, value=batch.product.pk)
        else:
            ws.cell(row=row_num, column=2, value="")

        # Column 2: Product Name
        if batch.product:
            ws.cell(row=row_num, column=1, value=str(batch.product))
        else:
            ws.cell(row=row_num, column=1, value="No product")

        # Column 3: Price
        ws.cell(row=row_num, column=3, value=float(batch.price))

        # Column 4: Quantity
        ws.cell(row=row_num, column=4, value=batch.quantity)

        # Column 5: Arrival Date
        arrival_date = batch.arrival_date
        if arrival_date:
            if hasattr(arrival_date, 'replace') and arrival_date.tzinfo is not None:
                arrival_date = arrival_date.replace(tzinfo=None)
            elif hasattr(arrival_date, 'date') and not hasattr(arrival_date, 'time'):
                arrival_date = datetime.combine(arrival_date, datetime.min.time())
        ws.cell(row=row_num, column=5, value=arrival_date)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'] = f'attachment; filename="warehouse_export_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'

    wb.save(response)
    return response